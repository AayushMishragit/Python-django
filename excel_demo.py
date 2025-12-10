from openpyxl import Workbook ,load_workbook

wb=Workbook()
ws=wb.active
ws['A1'] = "Name"
ws['B1'] = "Age"
ws['A2'] = "Alice"
ws['B2'] = 25


wb.save("output.xlsx")